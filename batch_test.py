#!/usr/bin/env python3
"""
批量测试脚本：对所有模型和数据集组合进行测试，并汇总结果到 Excel
"""
import os
import json
import subprocess
import tempfile
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import time

# 模型列表
MODELS = [
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen7grpo",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen7rflux",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen7srpo2",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen7srpo3",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen1.5grpo",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen1.5rflux",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen1.5srpo2",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/trained_models/qwen1.5srpo3",
]

# 数据集列表
DATASETS = [
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/data/aime2023.parquet",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/data/aime2024.parquet",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/data/aime2025.parquet",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/data/amc23.parquet",
    "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/data/math500split_val.parquet",
]

# 输出 Excel 文件路径
OUTPUT_EXCEL = "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/batch_evaluation_results.xlsx"

# 测试指标列表
METRICS = ["pass@1", "pass@8", "pass@16", "pass@32", "pass@64", "pass@128", "pass@256", "mean32"]


def extract_model_name(model_path):
    """从模型路径中提取模型名称"""
    return os.path.basename(model_path.rstrip('/'))


def extract_dataset_name(dataset_path):
    """从数据集路径中提取数据集名称"""
    filename = os.path.basename(dataset_path)
    return os.path.splitext(filename)[0]


def run_test(model_path, dataset_path):
    """运行单个测试并返回结果"""
    print(f"\n{'='*80}")
    print(f"测试: {extract_model_name(model_path)} - {extract_dataset_name(dataset_path)}")
    print(f"{'='*80}")
    
    # 创建临时 JSON 文件
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json_path = f.name
    
    try:
        # 运行测试
        cmd = [
            "python", "/mnt/shared-storage-user/mineru4s/dingruiyi/srpo/naivetest_vllm.py",
            "--model_path", model_path,
            "--dataset_path", dataset_path,
            "--test_mode", "all",
            "--output_json", json_path,
        ]

        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"错误: 测试失败")
            print(f"STDOUT: {result.stdout}")
            print(f"STDERR: {result.stderr}")
            return None
        
        # 读取结果
        if os.path.exists(json_path):
            with open(json_path, 'r') as f:
                data = json.load(f)
                return data.get('results', {})
        else:
            print(f"警告: JSON 文件不存在: {json_path}")
            return None
            
    except Exception as e:
        print(f"异常: {e}")
        return None
    finally:
        # 清理临时文件
        if os.path.exists(json_path):
            os.remove(json_path)


def create_excel(results_data):
    """创建 Excel 文件"""
    wb = Workbook()
    
    # 删除默认 sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 为每个数据集创建一个 sheet
    for dataset_path in DATASETS:
        dataset_name = extract_dataset_name(dataset_path)
        ws = wb.create_sheet(title=dataset_name)
        
        # 创建表头
        headers = ['Model'] + METRICS
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # 填充数据
        for model_path in MODELS:
            model_name = extract_model_name(model_path)
            key = f"{model_path}|||{dataset_path}"
            
            row = [model_name]
            if key in results_data:
                for metric in METRICS:
                    if metric in results_data[key]:
                        accuracy = results_data[key][metric]['accuracy']
                        row.append(accuracy)
                    else:
                        row.append(None)
            else:
                row.extend([None] * len(METRICS))
            
            ws.append(row)
        
        # 设置数据格式（百分比）
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(2, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    cell.number_format = '0.00%'
        
        # 自动调整列宽
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column = get_column_letter(col_idx)
            for row in ws[column]:
                try:
                    if row.value:
                        max_length = max(max_length, len(str(row.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column].width = adjusted_width
    
    # 保存文件
    wb.save(OUTPUT_EXCEL)
    print(f"\n{'='*80}")
    print(f"Excel 文件已保存: {OUTPUT_EXCEL}")
    print(f"{'='*80}")


def main():
    """主函数"""
    print("="*80)
    print("批量测试开始")
    print("="*80)
    print(f"模型数量: {len(MODELS)}")
    print(f"数据集数量: {len(DATASETS)}")
    print(f"总测试数: {len(MODELS) * len(DATASETS)}")
    print("="*80)
    
    results_data = {}
    total_tests = len(MODELS) * len(DATASETS)
    current_test = 0
    
    start_time = time.time()
    
    for model_path in MODELS:
        for dataset_path in DATASETS:
            current_test += 1
            print(f"\n进度: {current_test}/{total_tests}")
            
            key = f"{model_path}|||{dataset_path}"
            results = run_test(model_path, dataset_path)
            
            if results:
                results_data[key] = results
                print(f"✓ 测试完成")
            else:
                print(f"✗ 测试失败")
            
            # 每完成一个测试就保存一次（防止中途中断丢失数据）
            if results_data:
                create_excel(results_data)
    
    end_time = time.time()
    elapsed_time = end_time - start_time
    
    print("\n" + "="*80)
    print("批量测试完成")
    print("="*80)
    print(f"总耗时: {elapsed_time/3600:.2f} 小时 ({elapsed_time/60:.2f} 分钟)")
    print(f"成功: {len(results_data)}/{total_tests}")
    print(f"Excel 文件: {OUTPUT_EXCEL}")
    print("="*80)


if __name__ == "__main__":
    main()

